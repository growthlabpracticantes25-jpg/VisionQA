import os
import csv
from datetime import datetime
import requests
import streamlit as st
import cv2
import streamlit.components.v1 as components
from textwrap import dedent
import matplotlib.pyplot as plt
import pandas as pd
from io import BytesIO
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
)
from openpyxl.utils import get_column_letter
from modelo_ia import clasificar_imagen
from gemini_analisis import analizar_causas
from streamlit_option_menu import option_menu
import io
import re
from datetime import datetime
from html import escape
import base64
from pathlib import Path
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)


def cargar_css(nombre_archivo):
    ruta = Path(__file__).parent / "assets" / "css" / nombre_archivo

    if ruta.exists():
        with open(ruta, "r", encoding="utf-8") as archivo:
            st.markdown(f"<style>{archivo.read()}</style>", unsafe_allow_html=True)
    else:
        st.warning(f"No se encontró el archivo CSS: {ruta}")


# ---------------- APP PRINCIPAL ----------------

st.set_page_config(page_title="VisionQA", page_icon="🔍", layout="wide")

st.markdown(
    """
    <style>

    /* Fondo general */
    .stApp {
        background-color: #f4f7fb;
    }

    /* Ocultar elementos predeterminados de Streamlit */
    #MainMenu {
        visibility: hidden;
    }

    footer {
        visibility: hidden;
    }

    header {
        visibility: hidden;
    }

    /* Espacio del contenido principal */
    .block-container {
        padding-top: 1rem;
        padding-bottom: 2rem;
        max-width: 100%;
    }

    /* Barra superior del Dashboard */
    .visionqa-header {
    background: linear-gradient(90deg, #168db5, #24a0c1);
    border-radius: 10px;
    padding: 18px 24px;
    margin-bottom: 22px;
    color: white;
    box-shadow: 0 4px 14px rgba(0, 0, 0, 0.12);
    display: flex;
    justify-content: space-between;
    align-items: center;
}

.visionqa-header-left {
    display: flex;
    flex-direction: column;
}

.visionqa-header-right {
    text-align: right;
    color: white;
}

.visionqa-saludo {
    color: white;
    font-size: 20px;
    font-weight: 700;
}

.visionqa-subtitulo {
    color: white;
    font-size: 13px;
    margin-top: 5px;
    opacity: 0.92;
}

.visionqa-fecha {
    color: white;
    font-size: 13px;
}

.visionqa-hora {
    color: white;
    font-size: 18px;
    font-weight: 700;
    margin-top: 2px;
}
    </style>
    """,
    unsafe_allow_html=True,
)
# ---------------- ESTILOS ----------------

with open("styles/styles.css", encoding="utf-8") as f:
    st.markdown(
        f"<style>{f.read()}</style>",
        unsafe_allow_html=True,
    )

st.markdown(
    """
    <link
        rel="stylesheet"
        href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css"
    >
    """,
    unsafe_allow_html=True,
)
# ---------------- VARIABLES GLOBALES ----------------

archivo_csv = "registro_inspecciones.csv"

os.makedirs("inspecciones", exist_ok=True)

if "inspeccion" not in st.session_state:
    st.session_state.inspeccion = False


# ---------------- FUNCIONES ----------------
def guardar_inspeccion_api(resultado, defecto, confianza, archivo, origen):
    url_api = "http://127.0.0.1:8000/api/inspecciones/"

    datos = {
        "resultado": resultado,
        "defecto": defecto,
        "confianza": float(confianza),
        "archivo": archivo,
        "origen": origen,
    }

    try:
        respuesta = requests.post(url_api, json=datos, timeout=10)

        if respuesta.status_code == 201:
            return True, "Inspección guardada en Django."

        return False, f"Error de API {respuesta.status_code}: {respuesta.text}"

    except requests.exceptions.ConnectionError:
        return False, "No se pudo conectar con Django."

    except requests.exceptions.RequestException as error:
        return False, f"Error al enviar la inspección: {error}"


def obtener_inspecciones_api():

    url_api = "http://127.0.0.1:8000/api/inspecciones/"

    try:

        respuesta = requests.get(url_api, timeout=10)

        if respuesta.status_code == 200:

            return respuesta.json()

    except requests.exceptions.RequestException:

        pass

    return []

def mostrar_titulo(icono, titulo, descripcion):
    html_titulo = (
        '<div style="margin-top:30px; margin-bottom:22px;">'
            '<div style="'
                'display:flex;'
                'align-items:center;'
                'gap:10px;'
                'color:#231F20;'
                'font-size:28px;'
                'font-weight:700;'
            '">'
                f'{icono_svg(icono, 30, 0)}'
                f'<span>{titulo}</span>'
            '</div>'
            '<div style="'
                'margin-top:6px;'
                'color:#64748B;'
                'font-size:14px;'
            '">'
                f'{descripcion}'
            '</div>'
        '</div>'
    )

    st.markdown(
        html_titulo,
        unsafe_allow_html=True,
    )

def mostrar_encabezado_seccion(titulo, descripcion=""):

    html = (
        '<div class="section-header">'
        f'<div class="section-title">{titulo}</div>'
        f'<div class="section-description">{descripcion}</div>'
        "</div>"
    )

    st.markdown(html, unsafe_allow_html=True)


def mostrar_estado_sistema():
    st.subheader("🟢 Estado del Sistema")

    col1, col2, col3 = st.columns(3)

    with col1:
        with st.container(border=True):
            st.markdown("""
    ### 🤖 Modelo IA

    **Estado:** Conectado

    🟢 Cargado correctamente
    """)

    with col2:
        with st.container(border=True):
            st.markdown("""
    ### 🧠 Gemini

    **Estado:** Disponible

    🟢 Conectado
    """)

    with col3:
        with st.container(border=True):
            st.markdown("""
    ### 💾 Base de datos

    **Estado:** Disponible

    🟢 Registro listo
    """)

    st.divider()


def cargar_datos_registro():

    url_api = "http://127.0.0.1:8000/api/inspecciones/"

    try:

        respuesta = requests.get(url_api, timeout=10)

        if respuesta.status_code != 200:

            return 0, 0, 0

        registros = respuesta.json()

        total = len(registros)
        aptas = 0
        no_aptas = 0

        for registro in registros:

            resultado = str(registro.get("resultado", "")).strip().upper()

            if resultado in ["APTO", "BUENA"]:

                aptas += 1

            elif resultado in ["NO APTO", "MALA"]:

                no_aptas += 1

        return total, aptas, no_aptas

    except requests.exceptions.RequestException:

        return 0, 0, 0


def mostrar_modulo_inspeccion():

    mostrar_encabezado_seccion(
        "Método de inspección",
        "Selecciona cómo deseas capturar la pieza para iniciar el análisis.",
    )

    if "metodo_inspeccion" not in st.session_state:
        st.session_state["metodo_inspeccion"] = None

    # Primero se crean las columnas
    col_metodo_1, col_metodo_2 = st.columns(2)

    # Después se usan
    with col_metodo_1:
        with st.container(border=True):

            st.markdown(
                """
                <div style="text-align:center; padding:15px;">
                    <div style="font-size:50px;">📂</div>
                    <h4 style="margin:6px 0;">Cargar imagen</h4>
                    <p style="color:#6c757d; margin:0;">
                        Selecciona una imagen desde tu computadora.
                    </p>
                </div>
                """,
                unsafe_allow_html=True,
            )

            carga_activa = st.session_state["metodo_inspeccion"] == "Cargar imagen"

            if st.button(
                "✓ Imagen seleccionada" if carga_activa else "Seleccionar imagen",
                key="seleccionar_carga",
                use_container_width=True,
                type="primary" if carga_activa else "secondary",
            ):
                st.session_state["metodo_inspeccion"] = "Cargar imagen"
                st.rerun()

    with col_metodo_2:
        with st.container(border=True):

            st.markdown(
                """
                <div style="text-align:center; padding:15px;">
                    <div style="font-size:50px;">📷</div>
                    <h4 style="margin:6px 0;">Tomar fotografía</h4>
                    <p style="color:#6c757d; margin:0;">
                        Captura una imagen utilizando la cámara.
                    </p>
                </div>
                """,
                unsafe_allow_html=True,
            )

            camara_activa = st.session_state["metodo_inspeccion"] == "Tomar fotografía"

            if st.button(
                "✓ Cámara seleccionada" if camara_activa else "Abrir cámara",
                key="seleccionar_camara",
                use_container_width=True,
                type="primary" if camara_activa else "secondary",
            ):
                st.session_state["metodo_inspeccion"] = "Tomar fotografía"
                st.rerun()

    opcion = st.session_state["metodo_inspeccion"]

    if opcion == "Cargar imagen":
        st.success("📂 Método seleccionado: Cargar imagen")

    elif opcion == "Tomar fotografía":
        st.success("📷 Método seleccionado: Tomar fotografía")

    else:
        st.info("👆 Selecciona un método de inspección para comenzar.")
    # -------- PROCESAR Y REGISTRAR INSPECCIÓN --------

    def procesar_inspeccion(imagen, nombre_archivo, origen):

        os.makedirs("inspecciones", exist_ok=True)

        ruta_imagen = os.path.join("inspecciones", nombre_archivo)
        with open(ruta_imagen, "wb") as archivo:

            archivo.write(imagen.getbuffer())

        with st.spinner("La inteligencia artificial está analizando la pieza..."):

            respuesta_modelo = clasificar_imagen(ruta_imagen)

        # -------- INTERPRETAR RESPUESTA DEL MODELO --------

        if isinstance(respuesta_modelo, dict):

            resultado = respuesta_modelo.get("estado", "DESCONOCIDO")

            confianza = respuesta_modelo.get("confianza", 0.0)

            defecto = respuesta_modelo.get("defecto", None)

            resultado_yolo = respuesta_modelo.get("resultado_yolo")

            if resultado_yolo is not None:
                imagen_resultado = resultado_yolo.plot()
            else:
                imagen_resultado = None

        else:

            st.error("No fue posible interpretar la respuesta del modelo.")
            return

        # -------- MOSTRAR RESULTADO --------

        mostrar_encabezado_seccion(
            "Resultado de la inspección",
            "Clasificación y nivel de confianza obtenido por el modelo.",
        )

        resultado_normalizado = str(resultado).strip().upper()

        if resultado_normalizado in ["APTO", "BUENA"]:

            resultado_registro = "APTO"
            st.success("✅ PIEZA APTA — La pieza cumple con los criterios de calidad.")

        elif resultado_normalizado in ["NO APTO", "MALA"]:

            resultado_registro = "NO APTO"
            st.error("❌ PIEZA NO APTA — La pieza requiere revisión o rechazo.")

            if defecto:
                st.markdown(
                    f"**Defecto detectado:** "
                    f"{str(defecto).replace('_', ' ').title()}"
                )

        else:

            resultado_registro = resultado_normalizado
            st.warning(f"⚠ Resultado: {resultado_normalizado}")

        try:
            confianza_numero = float(confianza)
        except (TypeError, ValueError):
            confianza_numero = 0.0

        # Si la confianza viene entre 0 y 1,
        # se convierte a porcentaje
        if confianza_numero <= 1:
            confianza_porcentaje = confianza_numero * 100
        else:
            confianza_porcentaje = confianza_numero

        # -------- GUARDAR EN DJANGO --------

        guardado_api, mensaje_api = guardar_inspeccion_api(
            resultado=resultado_registro,
            defecto=defecto or "",
            confianza=confianza_porcentaje,
            archivo=ruta_imagen,
            origen=origen,
        )

        if guardado_api:
            st.caption(f"✅ {mensaje_api}")
        else:
            st.warning(mensaje_api)

            # -------- PANEL DE RESULTADOS --------

        col_imagen, col_info = st.columns([2, 1])

        with col_imagen:
            st.markdown("### 🖼️ Imagen procesada")

            if imagen_resultado is not None:
                try:
                    imagen_rgb = cv2.cvtColor(imagen_resultado, cv2.COLOR_BGR2RGB)

                    st.image(imagen_rgb, width=600)

                except Exception:
                    st.image(imagen_resultado, width=600)

        with col_info:

            with st.container(border=True):

                st.markdown("### 📊 Confianza del modelo")

                st.metric(label="Resultado", value=f"{confianza_porcentaje:.2f}%")

                st.progress(min(max(confianza_porcentaje / 100, 0.0), 1.0))

            with st.container(border=True):

                st.markdown("### 📂 Origen de la imagen")

                st.metric(label="Fuente", value=origen)

            if defecto:

                with st.container(border=True):

                    defecto_texto = str(defecto).replace("_", " ").title()

                    st.markdown("### ⚠️ Defecto detectado")

                    st.metric(label="Clasificación", value=defecto_texto)

                    # -------- GUARDAR EN CSV --------

            fecha_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

            nueva_fila = [
                fecha_hora,
                resultado_registro,
                defecto or "",
                f"{confianza_porcentaje:.2f}",
                nombre_archivo,
                origen,
            ]

            with open(archivo_csv, mode="a", newline="", encoding="utf-8") as archivo:

                escritor = csv.writer(archivo)
                escritor.writerow(nueva_fila)
        # -------- CARGAR IMAGEN --------

    if opcion == "Cargar imagen":

        archivo_subido = st.file_uploader(
            "Selecciona una imagen de la pieza",
            type=["jpg", "jpeg", "png"],
            key="imagen_cargada",
        )

        if archivo_subido is not None:

            analizar = st.button(
                "🔍 Analizar imagen",
                key="boton_analizar_archivo",
                use_container_width=True,
            )

            if analizar:

                nombre_archivo = datetime.now().strftime("archivo_%Y%m%d_%H%M%S.jpg")

                procesar_inspeccion(archivo_subido, nombre_archivo, "Archivo local")

            else:

                st.image(
                    archivo_subido,
                    caption="Imagen seleccionada",
                    use_container_width=True,
                )
    # -------- TOMAR FOTOGRAFÍA --------

    elif opcion == "Tomar fotografía":

        fotografia = st.camera_input(
            "Coloca la pieza frente a la cámara", key="fotografia_camara"
        )

        if fotografia is not None:

            if st.button(
                "📷 Analizar fotografía",
                key="boton_analizar_camara",
                use_container_width=True,
            ):

                nombre_archivo = datetime.now().strftime("inspeccion_%Y%m%d_%H%M%S.jpg")

                procesar_inspeccion(fotografia, nombre_archivo, "Cámara")

    st.divider()


def mostrar_resumen(total, aptas, no_aptas):

    mostrar_encabezado_seccion(
        "Resumen general", "Indicadores principales de las inspecciones registradas."
    )

    porcentaje_aptas = (aptas / total) * 100 if total > 0 else 0

    porcentaje_no_aptas = (no_aptas / total) * 100 if total > 0 else 0

    col1, col2, col3 = st.columns(3)

    tarjetas = [
        {
            "icono": "bi bi-clipboard-data",
            "valor": total,
            "titulo": "Total de inspecciones",
            "descripcion": "Registros almacenados",
            "detalle": "Base de datos actualizada",
            "clase": "kpi-total",
        },
        {
            "icono": "bi bi-check-circle",
            "valor": aptas,
            "titulo": "Piezas aptas",
            "descripcion": "Cumplen con calidad",
            "detalle": f"{porcentaje_aptas:.1f}% del total",
            "clase": "kpi-success",
        },
        {
            "icono": "bi bi-exclamation-triangle",
            "valor": no_aptas,
            "titulo": "Piezas no aptas",
            "descripcion": "Requieren revisión",
            "detalle": f"{porcentaje_no_aptas:.1f}% del total",
            "clase": "kpi-danger",
        },
    ]

    for columna, tarjeta in zip([col1, col2, col3], tarjetas):

        with columna:

            html = (
                f'<div class="kpi-card {tarjeta["clase"]}">'
                '<div class="kpi-card-top">'
                '<div class="kpi-icon">'
                f'<i class="{tarjeta["icono"]}"></i>'
                "</div>"
                '<div class="kpi-detail">'
                f'{tarjeta["detalle"]}'
                "</div>"
                "</div>"
                '<div class="kpi-value">'
                f'{tarjeta["valor"]}'
                "</div>"
                '<div class="kpi-title">'
                f'{tarjeta["titulo"]}'
                "</div>"
                '<div class="kpi-description">'
                f'{tarjeta["descripcion"]}'
                "</div>"
                "</div>"
            )

            st.markdown(html, unsafe_allow_html=True)

    st.caption("Actualización automática basada en las inspecciones registradas.")

    st.divider()


def mostrar_graficas(aptas, no_aptas):

    mostrar_encabezado_seccion(
        "Análisis de inspección", "Comparación visual entre piezas aptas y no aptas."
    )

    col1, col2 = st.columns(2)

    etiquetas = ["Aptas", "No aptas"]
    valores = [aptas, no_aptas]

    with col1:

        st.markdown("### Resultados de inspección")

        fig_barras, ax_barras = plt.subplots()

        barras = ax_barras.bar(etiquetas, valores)

        ax_barras.set_ylabel("Cantidad")
        ax_barras.set_title("Piezas inspeccionadas")

        for barra, valor in zip(barras, valores):

            ax_barras.text(
                barra.get_x() + barra.get_width() / 2,
                barra.get_height(),
                str(valor),
                ha="center",
                va="bottom",
            )

        st.pyplot(fig_barras)

    with col2:

        st.markdown("### Distribución de resultados")

        total = aptas + no_aptas

        if total > 0:

            fig_dona, ax_dona = plt.subplots()

            ax_dona.pie(
                valores,
                labels=etiquetas,
                autopct="%1.1f%%",
                startangle=90,
                wedgeprops={"width": 0.40},
            )

            ax_dona.axis("equal")

            st.pyplot(fig_dona)

        else:

            st.info("Todavía no hay inspecciones para mostrar la distribución.")

    st.divider()


def mostrar_indicadores(aptas, no_aptas):

    inspecciones_validas = aptas + no_aptas

    if inspecciones_validas == 0:
        st.info("Todavía no hay inspecciones suficientes para calcular indicadores.")
        st.divider()
        return

    porcentaje_apto = (aptas / inspecciones_validas) * 100

    porcentaje_no_apto = (no_aptas / inspecciones_validas) * 100

    mostrar_encabezado_seccion(
        "Indicadores de calidad", "Métricas porcentuales del desempeño del proceso."
    )

    col4, col5 = st.columns(2)

    with col4:
        st.markdown(
            f"""
            <div class="kpi-card">
                <div class="kpi-icon">📈</div>
                <div class="kpi-value">{porcentaje_apto:.1f}%</div>
                <div class="kpi-label">Tasa de aprobación</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col5:
        st.markdown(
            f"""
            <div class="kpi-card">
                <div class="kpi-icon">📉</div>
                <div class="kpi-value">{porcentaje_no_apto:.1f}%</div>
                <div class="kpi-label">Tasa de rechazo</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.divider()


def mostrar_registro():

    url_api = "http://127.0.0.1:8000/api/inspecciones/"

    try:

        respuesta = requests.get(url_api, timeout=10)

        if respuesta.status_code != 200:

            st.error("No fue posible consultar el registro en Django.")

            st.caption(f"Error de API: {respuesta.status_code}")

            st.divider()
            return

        registros = respuesta.json()

        if not registros:

            st.info("Todavía no hay inspecciones registradas.")

            st.divider()
            return

        datos = pd.DataFrame(registros)

        # Cambiar nombres de columnas
        datos = datos.rename(
            columns={
                "fecha": "Fecha",
                "resultado": "Resultado",
                "defecto": "Defecto",
                "confianza": "Confianza (%)",
                "archivo": "Archivo Guardado",
                "origen": "Origen",
            }
        )

        # Convertir y ordenar fechas
        datos["Fecha"] = pd.to_datetime(datos["Fecha"], errors="coerce")

        datos = datos.sort_values(by="Fecha", ascending=True).reset_index(drop=True)

        datos["Fecha"] = datos["Fecha"].dt.strftime("%d/%m/%Y %H:%M:%S")

        # -------- ÚLTIMA INSPECCIÓN --------

        ultima_inspeccion = datos.iloc[-1]

        st.markdown("### Última inspección")

        resultado = str(ultima_inspeccion["Resultado"]).strip().upper()

        defecto = ultima_inspeccion.get("Defecto", "")

        if pd.notna(defecto) and str(defecto).strip():

            defecto_texto = str(defecto).replace("_", " ").title()

        else:

            defecto_texto = "Sin defecto"

        confianza = float(ultima_inspeccion["Confianza (%)"])

        col_resultado, col_defecto, col_confianza = st.columns(3)

        # -------- TARJETA RESULTADO --------

        with col_resultado:

            with st.container(border=True):

                st.markdown("### 📋 Resultado")

                if resultado in ["APTO", "BUENA"]:

                    st.success("✅ PIEZA APTA")

                elif resultado in ["NO APTO", "MALA"]:

                    st.error("❌ PIEZA NO APTA")

                else:

                    st.warning(f"⚠ {resultado}")

                st.caption("Clasificación general de la última inspección.")

        # -------- TARJETA DEFECTO --------

        with col_defecto:

            with st.container(border=True):

                st.markdown("### ⚠️ Defecto")

                st.metric(
                    label="Clasificación detectada",
                    value=defecto_texto,
                )

                st.caption("Tipo de defecto identificado por el modelo.")

        # -------- TARJETA CONFIANZA --------

        with col_confianza:

            with st.container(border=True):

                st.markdown("### 🛡️ Confianza")

                st.metric(
                    label="Nivel del modelo",
                    value=f"{confianza:.2f}%",
                )

                st.progress(
                    min(
                        max(confianza / 100, 0.0),
                        1.0,
                    )
                )

        # -------- FECHA Y ORIGEN --------

        with st.container(border=True):

            col_fecha, col_origen = st.columns(2)

            with col_fecha:

                st.markdown(f"📅 **Fecha:** {ultima_inspeccion['Fecha']}")

            with col_origen:

                st.markdown(f"📂 **Origen:** {ultima_inspeccion['Origen']}")

        # -------- HISTORIAL COMPLETO --------

        st.markdown("### 📋 Historial de inspecciones")

        columnas_mostradas = [
            "Fecha",
            "Resultado",
            "Defecto",
            "Confianza (%)",
            "Origen",
        ]

        datos_mostrados = datos[columnas_mostradas].iloc[::-1].reset_index(drop=True)

        # Mejorar visualización del resultado
        datos_mostrados["Resultado"] = datos_mostrados["Resultado"].replace(
            {
                "APTO": "🟢 APTO",
                "BUENA": "🟢 APTO",
                "NO APTO": "🔴 NO APTO",
                "MALA": "🔴 NO APTO",
            }
        )

        # Mejorar visualización del defecto
        datos_mostrados["Defecto"] = (
            datos_mostrados["Defecto"]
            .fillna("Sin defecto")
            .astype(str)
            .str.replace("_", " ", regex=False)
            .str.title()
        )

        # Formato de confianza
        datos_mostrados["Confianza (%)"] = (
            pd.to_numeric(
                datos_mostrados["Confianza (%)"],
                errors="coerce",
            )
            .fillna(0)
            .round(2)
        )
        st.markdown(
            """
            <style>
            div[data-testid="stDataFrame"] table {
                border-radius: 12px;
                overflow: hidden;
            }

            div[data-testid="stDataFrame"] th {
                background: #0F6CBD;
                color: white;
                text-align: center;
                font-size: 15px;
            }

            div[data-testid="stDataFrame"] td {
                text-align: center;
                font-size: 14px;
            }
            </style>
            """,
            unsafe_allow_html=True,
        )

        st.dataframe(
            datos_mostrados,
            use_container_width=True,
            hide_index=True,
            height=420,
            column_config={
                "Fecha": st.column_config.TextColumn(
                    "📅 Fecha",
                    width="medium",
                ),
                "Resultado": st.column_config.TextColumn(
                    "Resultado",
                    width="small",
                ),
                "Defecto": st.column_config.TextColumn(
                    "⚠️ Defecto",
                    width="medium",
                ),
                "Confianza (%)": st.column_config.ProgressColumn(
                    "📊 Confianza",
                    min_value=0,
                    max_value=100,
                    format="%.2f%%",
                    width="medium",
                ),
                "Origen": st.column_config.TextColumn(
                    "📂 Origen",
                    width="small",
                ),
            },
        )
        st.info(f"📊 Total de inspecciones registradas: **{len(datos)}**")

    except requests.exceptions.ConnectionError:

        st.error("No se pudo conectar con Django.")

        st.info("Verifica que esté ejecutándose: " "python manage.py runserver")

    except requests.exceptions.RequestException as error:

        st.error("Ocurrió un error al consultar la API.")

        st.caption(f"Detalle técnico: {error}")

    except Exception as error:

        st.error("No fue posible cargar el registro de inspecciones.")

        st.caption(f"Detalle técnico: {error}")

    st.divider()


def generar_excel_registros(registros):

    datos = pd.DataFrame(registros)

    datos = datos.rename(
        columns={
            "fecha": "Fecha",
            "resultado": "Resultado",
            "defecto": "Defecto",
            "confianza": "Confianza (%)",
            "archivo": "Archivo Guardado",
            "origen": "Origen",
        }
    )

    columnas = [
        "Fecha",
        "Resultado",
        "Defecto",
        "Confianza (%)",
        "Archivo Guardado",
        "Origen",
    ]

    datos = datos[columnas]

    wb = Workbook()
    ws = wb.active
    ws.title = "Registro VisionQA"

    # -------- TÍTULO --------

    ws.merge_cells("A1:F1")
    ws["A1"] = "VisionQA - Historial de Inspecciones"

    ws["A1"].font = Font(
        size=16,
        bold=True,
        color="FFFFFF",
    )

    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A1"].fill = PatternFill(
        fill_type="solid",
        start_color="1F77B4",
        end_color="1F77B4",
    )

    # -------- SUBTÍTULO --------

    ws.merge_cells("A2:F2")

    ws["A2"] = (
        "Sistema Inteligente de Inspección Visual "
        "Asistido por Inteligencia Artificial"
    )

    ws["A2"].font = Font(
        italic=True,
        size=10,
    )

    ws["A2"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    # -------- INFORMACIÓN --------

    ws["A4"] = "Fecha de generación:"
    ws["B4"] = datetime.now().strftime("%d/%m/%Y %H:%M")

    ws["D4"] = "Registros:"
    ws["E4"] = len(datos)

    borde_delgado = Border(
        left=Side(style="thin", color="D9E1E8"),
        right=Side(style="thin", color="D9E1E8"),
        top=Side(style="thin", color="D9E1E8"),
        bottom=Side(style="thin", color="D9E1E8"),
    )

    # -------- ENCABEZADOS --------

    fila_encabezado = 6

    for numero_columna, nombre in enumerate(
        columnas,
        start=1,
    ):

        celda = ws.cell(
            row=fila_encabezado,
            column=numero_columna,
            value=nombre,
        )

        celda.font = Font(
            bold=True,
            color="FFFFFF",
        )

        celda.fill = PatternFill(
            fill_type="solid",
            start_color="1F7FB4",
            end_color="1F7FB4",
        )

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        celda.border = borde_delgado

    # -------- REGISTROS --------

    fila_actual = fila_encabezado + 1

    for fila in datos.itertuples(index=False):

        for numero_columna, valor in enumerate(
            fila,
            start=1,
        ):

            celda = ws.cell(
                row=fila_actual,
                column=numero_columna,
                value=valor,
            )

            celda.border = borde_delgado

            celda.alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )

            if numero_columna == 2:

                resultado = str(valor).strip().upper()

                if resultado == "APTO":

                    celda.fill = PatternFill(
                        fill_type="solid",
                        start_color="D9EAD3",
                        end_color="D9EAD3",
                    )

                    celda.font = Font(
                        bold=True,
                        color="008000",
                    )

                elif resultado == "NO APTO":

                    celda.fill = PatternFill(
                        fill_type="solid",
                        start_color="F4CCCC",
                        end_color="F4CCCC",
                    )

                    celda.font = Font(
                        bold=True,
                        color="C00000",
                    )

        fila_actual += 1

    # -------- AJUSTAR COLUMNAS --------

    from openpyxl.utils import get_column_letter

    for indice_columna in range(1, 7):

        letra = get_column_letter(indice_columna)

        longitud = 0

        for fila in ws.iter_rows(
            min_row=6,
            max_row=fila_actual - 1,
            min_col=indice_columna,
            max_col=indice_columna,
        ):

            celda = fila[0]

            if celda.value is not None:

                longitud = max(longitud, len(str(celda.value)))

        ws.column_dimensions[letra].width = min(
            longitud + 4,
            35,
        )

    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:F{fila_actual - 1}"

    archivo_excel = BytesIO()

    wb.save(archivo_excel)
    archivo_excel.seek(0)

    return archivo_excel.getvalue(), len(datos)


def limpiar_texto_pdf(texto):

    texto = str(texto)

    # Eliminar símbolos Markdown.
    texto = texto.replace("**", "")
    texto = texto.replace("__", "")
    texto = texto.replace("`", "")

    # Eliminar emojis y símbolos que Helvetica no puede mostrar.
    texto = re.sub(
        r"[\U00010000-\U0010ffff]",
        "",
        texto,
    )

    return texto.strip()


def generar_pdf_informe_ia(
    resultado_ia,
    estado,
    confianza,
    prioridad,
    defecto,
    accion,
    tipo_analisis,
):

    buffer = io.BytesIO()

    documento = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=1.7 * cm,
        leftMargin=1.7 * cm,
        topMargin=1.7 * cm,
        bottomMargin=1.7 * cm,
        title="Informe Ejecutivo VisionQA",
        author="VisionQA",
    )

    estilos = getSampleStyleSheet()

    estilo_titulo = ParagraphStyle(
        name="TituloVisionQA",
        parent=estilos["Title"],
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=24,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0F6CBD"),
        spaceAfter=8,
    )

    estilo_subtitulo = ParagraphStyle(
        name="SubtituloVisionQA",
        parent=estilos["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=13,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#667085"),
        spaceAfter=16,
    )

    estilo_seccion = ParagraphStyle(
        name="SeccionVisionQA",
        parent=estilos["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=18,
        textColor=colors.HexColor("#172033"),
        spaceBefore=12,
        spaceAfter=7,
    )

    estilo_subseccion = ParagraphStyle(
        name="SubseccionVisionQA",
        parent=estilos["Heading3"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=15,
        textColor=colors.HexColor("#29384D"),
        spaceBefore=8,
        spaceAfter=4,
    )

    estilo_texto = ParagraphStyle(
        name="TextoVisionQA",
        parent=estilos["BodyText"],
        fontName="Helvetica",
        fontSize=9.5,
        leading=14,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#303846"),
        spaceAfter=5,
    )

    estilo_vineta = ParagraphStyle(
        name="VinetaVisionQA",
        parent=estilo_texto,
        leftIndent=12,
        firstLineIndent=-7,
        bulletIndent=3,
        spaceAfter=3,
    )

    elementos = []

    # ---------------------------------------------------------
    # ENCABEZADO
    # ---------------------------------------------------------

    elementos.append(
        Paragraph(
            "VisionQA - Informe Ejecutivo IA",
            estilo_titulo,
        )
    )

    elementos.append(
        Paragraph(
            "Sistema Inteligente de Inspeccion Visual Asistido "
            "por Inteligencia Artificial",
            estilo_subtitulo,
        )
    )

    informacion = [
        [
            Paragraph("<b>Tipo de analisis</b>", estilo_texto),
            Paragraph(
                escape(limpiar_texto_pdf(tipo_analisis)),
                estilo_texto,
            ),
        ],
        [
            Paragraph("<b>Fecha de generacion</b>", estilo_texto),
            Paragraph(
                datetime.now().strftime("%d/%m/%Y %H:%M"),
                estilo_texto,
            ),
        ],
    ]

    tabla_informacion = Table(
        informacion,
        colWidths=[4.2 * cm, 12.5 * cm],
    )

    tabla_informacion.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (0, -1),
                    colors.HexColor("#EAF3FA"),
                ),
                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.6,
                    colors.HexColor("#C9D5E2"),
                ),
                (
                    "INNERGRID",
                    (0, 0),
                    (-1, -1),
                    0.3,
                    colors.HexColor("#D9E1E8"),
                ),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    elementos.append(tabla_informacion)
    elementos.append(Spacer(1, 14))

    # ---------------------------------------------------------
    # INDICADORES
    # ---------------------------------------------------------

    elementos.append(
        Paragraph(
            "Resumen de la inspeccion",
            estilo_seccion,
        )
    )

    datos_indicadores = [
        [
            Paragraph("<b>Resultado</b>", estilo_texto),
            Paragraph("<b>Confianza</b>", estilo_texto),
            Paragraph("<b>Prioridad</b>", estilo_texto),
            Paragraph("<b>Defecto</b>", estilo_texto),
            Paragraph("<b>Accion</b>", estilo_texto),
        ],
        [
            Paragraph(
                escape(limpiar_texto_pdf(estado)),
                estilo_texto,
            ),
            Paragraph(
                escape(f"{confianza:.1f}%"),
                estilo_texto,
            ),
            Paragraph(
                escape(limpiar_texto_pdf(prioridad)),
                estilo_texto,
            ),
            Paragraph(
                escape(limpiar_texto_pdf(defecto)),
                estilo_texto,
            ),
            Paragraph(
                escape(limpiar_texto_pdf(accion)),
                estilo_texto,
            ),
        ],
    ]

    tabla_indicadores = Table(
        datos_indicadores,
        colWidths=[
            3.15 * cm,
            3.15 * cm,
            3.15 * cm,
            3.65 * cm,
            3.65 * cm,
        ],
    )

    tabla_indicadores.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#DCECF8"),
                ),
                (
                    "BACKGROUND",
                    (0, 1),
                    (-1, 1),
                    colors.white,
                ),
                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.7,
                    colors.HexColor("#BFCBD8"),
                ),
                (
                    "INNERGRID",
                    (0, 0),
                    (-1, -1),
                    0.4,
                    colors.HexColor("#D9E1E8"),
                ),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )

    elementos.append(tabla_indicadores)
    elementos.append(Spacer(1, 14))

    # ---------------------------------------------------------
    # CONTENIDO GENERADO POR GEMINI
    # ---------------------------------------------------------

    lineas = str(resultado_ia).splitlines()

    for linea in lineas:

        linea = limpiar_texto_pdf(linea)

        if not linea:
            continue

        if linea.startswith("# "):

            titulo = linea[2:].strip()

            elementos.append(
                Paragraph(
                    escape(titulo),
                    estilo_seccion,
                )
            )

        elif linea.startswith("## "):

            titulo = linea[3:].strip()

            elementos.append(
                Paragraph(
                    escape(titulo),
                    estilo_seccion,
                )
            )

        elif linea.startswith("### "):

            titulo = linea[4:].strip()

            elementos.append(
                Paragraph(
                    escape(titulo),
                    estilo_subseccion,
                )
            )

        elif linea.startswith("- "):

            contenido = linea[2:].strip()

            elementos.append(
                Paragraph(
                    f"- {escape(contenido)}",
                    estilo_vineta,
                )
            )

        elif re.match(r"^\d+\.\s", linea):

            elementos.append(
                Paragraph(
                    escape(linea),
                    estilo_vineta,
                )
            )

        elif linea == "---":

            elementos.append(Spacer(1, 5))

        else:

            elementos.append(
                Paragraph(
                    escape(linea),
                    estilo_texto,
                )
            )

    elementos.append(Spacer(1, 18))

    elementos.append(
        Paragraph(
            "Informe generado por VisionQA mediante Gemini IA. "
            "Proyecto desarrollado para IOT Technologies.",
            estilo_subtitulo,
        )
    )

    documento.build(elementos)

    buffer.seek(0)

    return buffer.getvalue()


def mostrar_gemini():

    # -------- VARIABLES DEL INFORME --------

    resultado_ia = None
    titulo_informe = None
    descripcion_informe = None
    mostrar_tarjetas = False

    # -------- VARIABLES DE LAS TARJETAS --------

    estado_tarjeta = "Sin datos"
    confianza_tarjeta = 0.0
    prioridad_tarjeta = "Sin definir"
    defecto_tarjeta = "Sin defecto"
    accion_tarjeta = "Sin recomendación"

    # =========================================================
    # PANTALLA INICIAL
    # =========================================================

    if resultado_ia is None:

        # -------- FUNCIONES DEL MÓDULO --------

        st.markdown("### ¿Qué puede hacer este módulo?")

        funcion1, funcion2, funcion3, funcion4 = st.columns(4)

        with funcion1:

            with st.container(border=True):

                st.info("🔎 Detectar patrones")

                st.write("Identifica defectos repetitivos y posibles tendencias.")

        with funcion2:

            with st.container(border=True):

                st.success("⚙️ Analizar causas")

                st.write("Organiza las causas potenciales mediante la metodología 6M.")

        with funcion3:

            with st.container(border=True):

                st.warning("⚡ Recomendar acciones")

                st.write("Propone acciones correctivas y oportunidades de mejora.")

        with funcion4:

            with st.container(border=True):

                st.info("📋 Generar informes")

                st.write("Presenta los resultados en un formato ejecutivo y claro.")

        # -------- PROCESO DEL ANÁLISIS --------

        st.markdown("### Proceso del análisis")

        paso1, flecha1, paso2, flecha2, paso3, flecha3, paso4 = st.columns(
            [3, 1, 3, 1, 3, 1, 3]
        )

        with paso1:

            with st.container(border=True):

                st.markdown("#### 1️⃣ Selección")

                st.caption("El usuario selecciona el tipo de análisis.")

        with flecha1:
            st.markdown("## ➜")

        with paso2:

            with st.container(border=True):

                st.markdown("#### 2️⃣ Consulta")

                st.caption("VisionQA obtiene los registros almacenados.")

        with flecha2:
            st.markdown("## ➜")

        with paso3:

            with st.container(border=True):

                st.markdown("#### 3️⃣ Análisis")

                st.caption("Gemini interpreta los resultados y posibles causas.")

        with flecha3:
            st.markdown("## ➜")

        with paso4:

            with st.container(border=True):

                st.markdown("#### 4️⃣ Decisión")

                st.caption("Se generan recomendaciones para el supervisor.")

        st.info("""
            💡 Selecciona el análisis individual para revisar un caso
            específico o el análisis histórico para identificar patrones.
            """)

        st.caption(
            "🔒 Los registros se utilizan únicamente para consulta. "
            "El análisis no modifica la información almacenada."
        )
        # =====================================================
        # SELECCIÓN DEL TIPO DE ANÁLISIS
        # =====================================================

        st.divider()

        st.markdown("## 🚀 Selecciona el tipo de análisis")

        st.caption("Elige la información que deseas analizar mediante IA Generativa.")

        opcion1, opcion2 = st.columns(2)

        # -----------------------------------------------------
        # ÚLTIMA INSPECCIÓN
        # -----------------------------------------------------

        with opcion1:

            with st.container(border=True):
                st.caption("🟢 ANÁLISIS INDIVIDUAL")
                st.markdown("## 🔎 Última inspección")

                st.write("""
                    Analiza únicamente el registro más reciente almacenado
                    en VisionQA y genera un diagnóstico individual.
                    """)

                st.success("✔ Identificación de posibles causas raíz")
                st.success("✔ Acciones correctivas priorizadas")
                st.success("✔ Recomendaciones mediante metodología 6M")

                st.caption("⏱️ Análisis individual del último registro disponible.")

                if st.button(
                    "🧠 Analizar última inspección",
                    key="analizar_ultima_inspeccion",
                    width="stretch",
                ):

                    registros = obtener_inspecciones_api()

                    if not registros:

                        st.warning("No hay inspecciones registradas.")

                    else:

                        ultima = registros[-1]

                        # -------- DATOS DE LAS TARJETAS --------

                        estado_tarjeta = (
                            str(ultima.get("resultado", "Sin datos")).strip().upper()
                        )

                        try:

                            confianza_tarjeta = float(ultima.get("confianza", 0))

                        except (TypeError, ValueError):

                            confianza_tarjeta = 0.0

                        defecto_tarjeta = (
                            str(ultima.get("defecto", "Sin defecto"))
                            .replace("_", " ")
                            .strip()
                            .title()
                        )

                        if not defecto_tarjeta:

                            defecto_tarjeta = "Sin defecto"

                        if estado_tarjeta in ["NO APTO", "MALA"]:

                            prioridad_tarjeta = "ALTA"
                            accion_tarjeta = "Contener pieza"

                        elif estado_tarjeta in ["APTO", "BUENA"]:

                            prioridad_tarjeta = "BAJA"
                            accion_tarjeta = "Liberar pieza"

                        else:

                            prioridad_tarjeta = "MEDIA"
                            accion_tarjeta = "Revisión manual"

                        # -------- DATOS PARA GEMINI --------

                        datos = f"""
Fecha: {ultima.get('fecha', 'Sin fecha')}
Resultado: {ultima.get('resultado', 'Sin resultado')}
Defecto: {ultima.get('defecto', 'Sin defecto')}
Confianza: {ultima.get('confianza', 0)}%
Origen: {ultima.get('origen', 'Sin origen')}
"""

                        with st.spinner(
                            "🧠 Gemini está analizando la última inspección..."
                        ):

                            resultado = analizar_causas(datos)

                        st.success("✅ Análisis completado correctamente.")

                        resultado_ia = resultado

                        titulo_informe = "🧠 Informe ejecutivo de la última inspección"

                        descripcion_informe = (
                            "Análisis basado en el registro más reciente "
                            "de VisionQA."
                        )

                        mostrar_tarjetas = True

        # -----------------------------------------------------
        # ÚLTIMAS 10 INSPECCIONES
        # -----------------------------------------------------

        with opcion2:

            with st.container(border=True):
                st.caption("🔵 ANÁLISIS HISTÓRICO")
                st.markdown("## 📊 Últimas 10 inspecciones")

                st.write("""
                    Analiza el historial reciente para identificar defectos
                    recurrentes, tendencias y oportunidades de mejora.
                    """)

                st.info("📈 Identificación de tendencias")
                st.info("📌 Detección de defectos recurrentes")
                st.info("📊 Análisis comparativo del historial")

                st.caption("⏱️ Análisis agrupado de hasta diez registros recientes.")

                if st.button(
                    "📊 Analizar historial",
                    key="analizar_ultimas_10_inspecciones",
                    width="stretch",
                ):

                    registros = obtener_inspecciones_api()

                    if not registros:

                        st.warning("No hay inspecciones registradas.")

                    else:

                        ultimas = registros[-10:]

                        texto = ""

                        for registro in ultimas:

                            texto += f"""
Fecha: {registro.get('fecha', 'Sin fecha')}
Resultado: {registro.get('resultado', 'Sin resultado')}
Defecto: {registro.get('defecto', 'Sin defecto')}
Confianza: {registro.get('confianza', 0)}%
Origen: {registro.get('origen', 'Sin origen')}

"""

                        with st.spinner(
                            "🧠 Gemini está analizando las últimas inspecciones..."
                        ):

                            resultado = analizar_causas(texto)

                        st.success("✅ Análisis completado correctamente.")

                        resultado_ia = resultado

                        titulo_informe = (
                            "📊 Informe ejecutivo de las últimas " "10 inspecciones"
                        )

                        descripcion_informe = (
                            "Análisis de patrones, defectos y oportunidades "
                            "de mejora."
                        )

                        mostrar_tarjetas = False

    # =========================================================
    # MOSTRAR INFORME A TODO LO ANCHO
    # =========================================================

    if resultado_ia is not None:
        # =====================================================
        # INFORME EJECUTIVO
        # =====================================================

        st.markdown(
            """
            <div style="
                background:linear-gradient(135deg,#0F6CBD,#1593B7);
                border-radius:18px;
                padding:24px;
                color:white;
                margin-bottom:25px;
            ">

            <div style="
                font-size:30px;
                font-weight:700;
                margin-bottom:8px;
            ">
                🧠 Informe Ejecutivo IA
            </div>

            <div style="
                font-size:15px;
                opacity:0.95;
                line-height:1.6;
            ">
                Análisis generado automáticamente mediante Gemini AI a partir de
                los registros de inspección de VisionQA para apoyar la toma de
                decisiones en calidad.
            </div>

            </div>
            """,
            unsafe_allow_html=True,
        )

        if mostrar_tarjetas:

            (
                col_resultado,
                col_confianza,
                col_prioridad,
                col_defecto,
                col_accion,
            ) = st.columns(5)

            # -------- RESULTADO --------

            with col_resultado:

                with st.container(border=True):

                    if estado_tarjeta in ["APTO", "BUENA"]:

                        st.success("✅ Resultado")

                    elif estado_tarjeta in ["NO APTO", "MALA"]:

                        st.error("❌ Resultado")

                    else:

                        st.warning("⚠️ Resultado")

                    st.metric(
                        label="Estado",
                        value=estado_tarjeta,
                    )

            # -------- CONFIANZA --------

            with col_confianza:

                with st.container(border=True):

                    st.info("📊 Confianza")

                    st.metric(
                        label="Nivel",
                        value=f"{confianza_tarjeta:.1f}%",
                    )

            # -------- PRIORIDAD --------

            with col_prioridad:

                with st.container(border=True):

                    if prioridad_tarjeta == "ALTA":

                        st.error("🔴 Prioridad")

                    elif prioridad_tarjeta == "MEDIA":

                        st.warning("🟡 Prioridad")

                    else:

                        st.success("🟢 Prioridad")

                    st.metric(
                        label="Nivel",
                        value=prioridad_tarjeta,
                    )

            # -------- DEFECTO --------

            with col_defecto:

                with st.container(border=True):

                    st.warning("🔎 Defecto")

                    st.metric(
                        label="Detectado",
                        value=defecto_tarjeta,
                    )

            # -------- ACCIÓN --------

            with col_accion:

                with st.container(border=True):

                    st.info("⚡ Acción")

                    st.metric(
                        label="Inmediata",
                        value=accion_tarjeta,
                    )

        # -------- INFORME GENERADO POR GEMINI --------

        resultado_limpio = (
            resultado_ia.replace("# 📄 Informe detallado", "")
            .replace("## 📄 Informe detallado", "")
            .replace("📄 Informe detallado", "")
            .replace(
                "Interpretación generada por IA utilizando la metodología "
                "Lean Manufacturing, Six Sigma y análisis 6M.",
                "",
            )
        )

        st.markdown(resultado_limpio.strip())
        # -------- GUARDAR INFORME PDF PARA REPORTES --------

        if mostrar_tarjetas:
            tipo_analisis_pdf = "Última inspección"
        else:
            tipo_analisis_pdf = "Últimas 10 inspecciones"

        pdf_informe = generar_pdf_informe_ia(
            resultado_ia=resultado_limpio,
            estado=estado_tarjeta,
            confianza=confianza_tarjeta,
            prioridad=prioridad_tarjeta,
            defecto=defecto_tarjeta,
            accion=accion_tarjeta,
            tipo_analisis=tipo_analisis_pdf,
        )

        nombre_pdf = (
            "VisionQA_Informe_IA_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".pdf"
        )

        st.session_state["pdf_informe_ia"] = pdf_informe
        st.session_state["nombre_pdf_ia"] = nombre_pdf
        st.session_state["fecha_informe_ia"] = datetime.now().strftime("%d/%m/%Y %H:%M")


def mostrar_reportes():

    st.markdown("## 📂 Documentos disponibles")

    st.caption("Selecciona el documento que deseas generar o descargar.")

    st.write("")

    col_excel, col_pdf = st.columns(2)

    # -------------------------------------------------
    # HISTORIAL
    # -------------------------------------------------

    with col_excel:

        with st.container(border=True):

            st.markdown("# 📊")

            st.markdown("### Historial de inspecciones")

            st.write("""
                Exporta todos los registros de inspección
                almacenados por VisionQA en formato Excel.
                """)

            st.info("Incluye fecha, resultado, confianza, " "defecto y origen.")

            registros = obtener_inspecciones_api()

            if not registros:

                st.warning("No existen inspecciones registradas para exportar.")

                st.button(
                    "📥 Excel no disponible",
                    key="excel_no_disponible",
                    width="stretch",
                    disabled=True,
                )

            else:

                try:

                    archivo_excel, total_registros = generar_excel_registros(registros)

                    nombre_excel = (
                        "Registro_VisionQA_"
                        + datetime.now().strftime("%Y%m%d_%H%M%S")
                        + ".xlsx"
                    )

                    st.download_button(
                        label="📥 Descargar Excel",
                        data=archivo_excel,
                        file_name=nombre_excel,
                        mime=(
                            "application/vnd.openxmlformats-"
                            "officedocument.spreadsheetml.sheet"
                        ),
                        key="descargar_excel_reportes",
                        width="stretch",
                    )

                    st.caption(f"Registros disponibles: {total_registros}")

                except Exception as error:

                    st.error("No fue posible preparar el archivo Excel.")

                    st.caption(str(error))
    # -------------------------------------------------
    # INFORME IA
    # -------------------------------------------------

    with col_pdf:

        with st.container(border=True):

            st.markdown("# 🤖")

            st.markdown("### Informe Ejecutivo IA")

            st.write("""
                Exporta el último análisis generado
                mediante Gemini IA en formato PDF.
                """)

            st.info("Incluye resumen, análisis 6M, " "acciones y recomendaciones.")

            if "pdf_informe_ia" in st.session_state:

                st.download_button(
                    label="📄 Descargar PDF",
                    data=st.session_state["pdf_informe_ia"],
                    file_name=st.session_state.get(
                        "nombre_pdf_ia",
                        "VisionQA_Informe_IA.pdf",
                    ),
                    mime="application/pdf",
                    key="descargar_pdf_reportes",
                    width="stretch",
                )

                st.caption(
                    "Último informe generado: "
                    + st.session_state.get(
                        "fecha_informe_ia",
                        "Fecha no disponible",
                    )
                )

            else:

                st.warning("Primero genera un análisis en la sección IA Generativa.")

                st.button(
                    "📄 PDF no disponible",
                    key="pdf_no_disponible",
                    width="stretch",
                    disabled=True,
                )

    st.divider()

    st.info("""
**Información**

• Los documentos se generan con la información almacenada en VisionQA.

• El historial de inspecciones se exporta en formato Excel (.xlsx).

• El informe ejecutivo se exporta en formato PDF (.pdf).
""")


def mostrar_footer():

    st.markdown(
        """
<div class="footer">
<b>VisionQA v1.0</b><br>
Sistema Inteligente de Inspección Visual Asistido por Inteligencia Artificial
<hr>
Instituto Superior de Ciencias de Ciudad Juárez<br>
Proyecto desarrollado para IOT Technologies
<hr>
© 2026 VisionQA
</div>
""",
        unsafe_allow_html=True,
    )


cargar_css("global.css")


def mostrar_login():

    st.markdown(
        """
        <style>
        [data-testid="stSidebar"] {
            display: none;
        }

        [data-testid="stHeader"] {
            background: transparent;
        }

        .block-container {
            max-width: 1180px;
            padding-top: 3rem;
            padding-bottom: 3rem;
        }

        .login-panel {
            min-height: 620px;
            border-radius: 28px;
            overflow: hidden;
            box-shadow: 0 20px 55px rgba(35, 31, 32, 0.16);
            background: #FFFFFF;
            border: 1px solid rgba(189, 198, 195, 0.55);
        }

        .login-brand {
            min-height: 620px;
            padding: 58px 48px;
            border-radius: 28px 0 0 28px;
            background:
                radial-gradient(
                    circle at 15% 20%,
                    rgba(152, 218, 233, 0.38),
                    transparent 28%
                ),
                radial-gradient(
                    circle at 90% 82%,
                    rgba(255, 255, 255, 0.14),
                    transparent 30%
                ),
                linear-gradient(
                    145deg,
                    #0032A0 0%,
                    #1D7EAE 58%,
                    #1998B7 100%
                );
            color: white;
        }

        .iot-mark {
            font-size: 16px;
            font-weight: 700;
            letter-spacing: 1.4px;
            margin-bottom: 95px;
        }

        .brand-symbol {
            width: 72px;
            height: 72px;
            border: 3px solid #98DAE9;
            border-radius: 50%;
            position: relative;
            margin-bottom: 28px;
        }

        .brand-symbol::before,
        .brand-symbol::after {
            content: "";
            position: absolute;
            border: 3px solid #98DAE9;
            border-radius: 50%;
        }

        .brand-symbol::before {
            inset: 10px;
        }

        .brand-symbol::after {
            inset: 22px;
            background: #98DAE9;
        }

        .brand-title {
            font-size: 46px;
            line-height: 1;
            font-weight: 800;
            margin-bottom: 18px;
        }

        .brand-subtitle {
            font-size: 21px;
            line-height: 1.45;
            font-weight: 500;
            max-width: 420px;
            margin-bottom: 22px;
        }

        .brand-copy {
            font-size: 15px;
            line-height: 1.65;
            max-width: 440px;
            color: rgba(255, 255, 255, 0.82);
        }

        .brand-badge {
            display: inline-block;
            margin-top: 38px;
            padding: 10px 16px;
            border-radius: 999px;
            border: 1px solid rgba(255, 255, 255, 0.30);
            background: rgba(255, 255, 255, 0.10);
            font-size: 13px;
        }

        .login-form-header {
            margin-top: 85px;
            margin-bottom: 28px;
        }

        .login-form-header h1 {
            color: #231F20;
            font-size: 38px;
            margin: 0 0 10px 0;
        }

        .login-form-header p {
            color: #667085;
            font-size: 15px;
            margin: 0;
        }

        div[data-testid="stTextInput"] label {
            color: #231F20;
            font-weight: 600;
            font-size: 14px;
        }

        div[data-testid="stTextInput"] input {
            border-radius: 14px;
            min-height: 48px;
            border: 1px solid #BDC6C3;
            background: #FFFFFF;
        }

        div[data-testid="stTextInput"] input:focus {
            border-color: #1D7EAE;
            box-shadow: 0 0 0 3px rgba(29, 126, 174, 0.13);
        }

        div[data-testid="stButton"] > button {
            width: 100%;
            min-height: 50px;
            border: none;
            border-radius: 14px;
            background: linear-gradient(
                90deg,
                #0032A0 0%,
                #1D7EAE 100%
            );
            color: white;
            font-weight: 700;
            font-size: 15px;
            box-shadow: 0 8px 20px rgba(0, 50, 160, 0.22);
        }

        div[data-testid="stButton"] > button:hover {
            border: none;
            color: white;
            background: linear-gradient(
                90deg,
                #1D7EAE 0%,
                #0032A0 100%
            );
        }

        .login-help {
            margin-top: 24px;
            padding-top: 20px;
            border-top: 1px solid #E8ECEB;
            color: #667085;
            font-size: 13px;
            line-height: 1.5;
        }

        .login-footer {
            margin-top: 42px;
            color: #98A2B3;
            font-size: 12px;
        }
        div[data-testid="stHorizontalBlock"]
        > div[data-testid="column"]:first-child {
            background: linear-gradient(
                145deg,
                #0032A0 0%,
                #1D7EAE 58%,
                #1998B7 100%
            ) !important;
 
            min-height: 620px;
            padding: 58px 48px;
            border-radius: 28px 0 0 28px;
        }

        div[data-testid="stHorizontalBlock"]
        > div[data-testid="column"]:nth-child(2) {
            min-height: 620px;
            padding: 30px 42px;
            background: #FFFFFF;
            border-radius: 0 28px 28px 0;
        }

        @media (max-width: 900px) {
            .login-brand {
                min-height: 420px;
                border-radius: 28px 28px 0 0;
                padding: 42px 32px;
            }

            .login-form-header {
                margin-top: 20px;
            }

            .iot-mark {
                margin-bottom: 45px;
            }

            .brand-title {
                font-size: 38px;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    columna_marca, columna_formulario = st.columns([1.15, 0.85], gap="large")

    with columna_marca:

        st.markdown("<p class='iot-mark'>IOT TECHNOLOGIES</p>", unsafe_allow_html=True)

        st.markdown("<div class='brand-symbol'></div>", unsafe_allow_html=True)

        st.markdown("<h1 class='brand-title'>VisionQA</h1>", unsafe_allow_html=True)

        st.markdown(
            """
            <p class="brand-subtitle">
                Sistema Inteligente de Inspección Visual
            </p>
            """,
            unsafe_allow_html=True,
        )

        st.markdown(
            """
            <p class="brand-copy">
                Plataforma de control de calidad orientada a la
                detección de defectos, gestión de inspecciones y
                análisis de resultados para procesos industriales.
            </p>
            """,
            unsafe_allow_html=True,
        )

        st.markdown(
            """
            <div class="brand-badge">
                Conectividad · Confiabilidad · Eficiencia
            </div>
            """,
            unsafe_allow_html=True,
        )

    with columna_formulario:

        st.markdown(
            """
            <div class="login-form-header">
                <h1>Iniciar sesión</h1>
                <p>
                    Ingresa tus credenciales para acceder a VisionQA.
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )

        usuario = st.text_input(
            "Usuario", placeholder="Escribe tu usuario", key="login_usuario"
        )

        contraseña = st.text_input(
            "Contraseña",
            type="password",
            placeholder="Escribe tu contraseña",
            key="login_contrasena",
        )

        iniciar = st.button("Iniciar sesión", key="login_boton")

        if iniciar:

            if not usuario.strip() or not contraseña:

                st.warning("Escribe el usuario y la contraseña.")

            else:

                try:

                    with st.spinner("Validando credenciales..."):

                        respuesta = requests.post(
                            "http://127.0.0.1:8000/api/login/",
                            json={"username": usuario.strip(), "password": contraseña},
                            timeout=10,
                        )

                    if respuesta.status_code == 200:

                        datos = respuesta.json()

                        if datos.get("success"):

                            st.session_state["logueado"] = True
                            st.session_state["usuario"] = datos.get("username", usuario)
                            st.session_state["is_staff"] = datos.get("is_staff", False)
                            st.session_state["is_superuser"] = datos.get(
                                "is_superuser", False
                            )

                            st.rerun()

                        else:

                            st.error("Usuario o contraseña incorrectos.")

                    else:

                        st.error("El servidor no pudo validar el acceso.")

                except requests.exceptions.ConnectionError:

                    st.error(
                        "No fue posible conectar con el servidor de "
                        "VisionQA. Verifica que Django esté ejecutándose."
                    )

                except requests.exceptions.Timeout:

                    st.error("El servidor tardó demasiado en responder.")

                except ValueError:

                    st.error("El servidor devolvió una respuesta no válida.")

                except Exception as error:

                    st.error(f"Ocurrió un error inesperado: {error}")

        st.markdown(
            """
            <div class="login-help">
                Acceso autorizado exclusivamente para personal
                registrado en el sistema.
            </div>

            <div class="login-footer">
                VisionQA v1.0 · IOT Technologies · 2026
            </div>
            """,
            unsafe_allow_html=True,
        )
def icono_svg(nombre, tamaño=22, margen_derecho=8):
    ruta = Path(__file__).parent / "assets" / "icons" / nombre

    if not ruta.exists():
        return ""

    contenido = base64.b64encode(
        ruta.read_bytes()
    ).decode("utf-8")

    return (
        f'<img '
        f'src="data:image/svg+xml;base64,{contenido}" '
        f'width="{tamaño}" '
        f'height="{tamaño}" '
        f'style="'
        f'vertical-align:middle;'
        f'margin-right:{margen_derecho}px;'
        f'object-fit:contain;'
        f'">'
    )

def main():
    if "logueado" not in st.session_state:
        st.session_state["logueado"] = True

    usuario_url = st.query_params.get("usuario")

    if usuario_url:
        st.session_state["usuario"] = usuario_url

    elif "usuario" not in st.session_state:
        st.session_state["usuario"] = "Usuario"

    # -------- MENÚ LATERAL --------
    ahora = datetime.now()

    dias_semana = [
        "Lunes",
        "Martes",
        "Miércoles",
        "Jueves",
        "Viernes",
        "Sábado",
        "Domingo",
    ]

    meses = [
        "Ene",
        "Feb",
        "Mar",
        "Abr",
        "May",
        "Jun",
        "Jul",
        "Ago",
        "Sep",
        "Oct",
        "Nov",
        "Dic",
    ]

    dia_semana = dias_semana[ahora.weekday()]
    mes_actual = meses[ahora.month - 1]

    fecha_actual = f"{dia_semana}, {ahora.day:02d} " f"{mes_actual} {ahora.year}"

    hora_actual = ahora.strftime("%I:%M %p")

    if ahora.hour < 12:
        saludo = "Buenos días"
    elif ahora.hour < 19:
        saludo = "Buenas tardes"
    else:
        saludo = "Buenas noches"

    nombre_usuario = escape(
        str(
            st.session_state.get(
                "usuario",
                "Usuario",
            )
        )
    )

    html_barra = (
        '<div class="visionqa-topbar">'

            '<div class="visionqa-topbar-left">'
                f'<div class="visionqa-topbar-greeting">'
                    f'{saludo}, {nombre_usuario}'
                '</div>'
                '<div class="visionqa-topbar-subtitle">'
                    'Este es el estado actual del sistema VisionQA'
                '</div>'
            '</div>'

            '<div class="visionqa-topbar-right">'

                '<div class="visionqa-topbar-datetime">'
                    f'<div class="visionqa-topbar-date">'
                        f'{fecha_actual}'
                    '</div>'
                    f'<div class="visionqa-topbar-time">'
                        f'{hora_actual}'
                    '</div>'
                '</div>'

                '<div class="visionqa-topbar-actions">'

                    f'<div class="visionqa-topbar-action" '
                    f'title="{nombre_usuario}">'
                        f'{icono_svg("user.svg",20,0)}'
                    '</div>'

                    '<a class="visionqa-topbar-action" '
                    'href="?pagina=Acerca%20de" '
                    'title="Ayuda">'
                        f'{icono_svg("help.svg",20,0)}'
                    '</a>'

                    '<a class="visionqa-topbar-action" '
                    'href="http://127.0.0.1:8000/api/logout/" '
                    'target="_self" '
                    'title="Cerrar sesión">'
                        f'{icono_svg("logout.svg",20,0)}'
                    '</a>'

                '</div>'

            '</div>'

        '</div>'
    )

    st.markdown(
        html_barra,
        unsafe_allow_html=True,
    )
    with st.sidebar:
        st.markdown( 
            f"""
            <div style="
                background:#0d6efd20;
                border:1px solid #0d6efd40;
                border-radius:10px;
                padding:10px;
                margin-bottom:15px;
                text-align:center;
            ">

            {icono_svg("user.svg",22,6)}

            <br>

            {st.session_state["usuario"]}

            </div>
            """,
            unsafe_allow_html=True,
        )

        st.image("assets/logo_iot.png", use_container_width=True)

        st.markdown(
            '<div class="sidebar-brand">'
            '<div class="sidebar-title">VisionQA</div>'
            '<div class="sidebar-subtitle">'
            "Sistema Inteligente de Inspección Visual"
            "</div>"
            "</div>",
            unsafe_allow_html=True,
        )

        st.markdown('<div class="sidebar-separator"></div>', unsafe_allow_html=True)

        pagina = option_menu(
            menu_title=None,
            options=[
                "Dashboard",
                "Inspección",
                "Registro",
                "IA Generativa",
                "Reportes",
                "Acerca de",
            ],
            icons=[
                "speedometer2",
                "search",
                "clipboard-data",
                "cpu",
                "file-earmark-bar-graph",
                "info-circle",
            ],
            menu_icon=None,
            default_index=0,
            orientation="vertical",
            styles={
                "container": {
                    "padding": "12px 10px",
                    "margin": "0px",
                    "background-color": "#1998B7",
                    "border": "none",
                    "border-radius": "0px",
                    "box-shadow": "none",
                },
                "icon": {"color": "#FFFFFF", "font-size": "18px"},
                "nav-link": {
                    "font-size": "15px",
                    "font-weight": "500",
                    "color": "#FFFFFF",
                    "text-align": "left",
                    "margin": "6px 0",
                    "padding": "14px 16px",
                    "border-radius": "10px",
                    "background-color": "#1998B7",
                    "border": "none",
                },
                "nav-link-hover": {
                    "background-color": "rgba(255,255,255,0.16)",
                    "color": "#FFFFFF",
                },
                "nav-link-selected": {
                    "background-color": "#0A4E95",
                    "color": "#FFFFFF",
                    "font-weight": "700",
                    "border-radius": "10px",
                    "box-shadow": "0 4px 10px rgba(0,0,0,0.18)",
                },
            },
        )
        st.markdown("---")

        st.markdown(
            """
                <div class="sidebar-footer">
                    <strong>VisionQA</strong><br>
                    Versión 1.0<br>
                    IOT Technologies
                </div>
                """,
            unsafe_allow_html=True,
        )

    # -------- BARRA SUPERIOR DEL USUARIO --------

    
    # -------- DASHBOARD --------
    if pagina == "Dashboard":
        registros = obtener_inspecciones_api()

        total = len(registros)

        aptas = sum(
            1
            for registro in registros
            if str(
                registro.get("resultado", "")
            ).strip().upper() == "APTO"
        )

        no_aptas = sum(
            1
            for registro in registros
            if str(
                registro.get("resultado", "")
            ).strip().upper() == "NO APTO"
        )
        st.markdown(
            f"""
            <h2 style="
                color:#231F20 !important;
                font-size:32px;
                font-weight:700;
                margin-top:0;
                margin-bottom:18px;
                display:flex;
                align-items:center;
            ">
                {icono_svg("dashboard.svg", 40, 12)}
                Dashboard Operativo
            </h2>
            """,
            unsafe_allow_html=True,
        )
        mostrar_resumen(total, aptas, no_aptas)

        mostrar_graficas(aptas, no_aptas)

        mostrar_indicadores(aptas, no_aptas)

    # -------- INSPECCIÓN --------

    elif pagina == "Inspección":

        mostrar_titulo(
            "inspection.svg",
            "Inspección Visual",
            "Captura y analiza piezas mediante inteligencia artificial.",
        )

        mostrar_estado_sistema()
        mostrar_modulo_inspeccion()

    # -------- REGISTRO --------

    elif pagina == "Registro":

        mostrar_titulo(
            "register.svg",
            "Registro de Inspecciones",
            "Consulta los resultados y el historial de inspecciones.",
        )

        mostrar_registro()
    # -------- IA GENERATIVA --------

    elif pagina == "IA Generativa":

        mostrar_titulo(
            "ai.svg",
            "Análisis Inteligente",
            "Analiza inspecciones mediante Gemini, metodología 6M, Lean Manufacturing y Six Sigma para apoyar la toma de decisiones en calidad.",
        )
        mostrar_gemini()

    # -------- REPORTES --------

    elif pagina == "Reportes":

        mostrar_titulo(
            "report.svg",
            "Centro de Reportes",
            "Genera y descarga los documentos disponibles del sistema VisionQA.",
        )

        mostrar_reportes()

    # -------- ACERCA DE --------

    elif pagina == "Acerca de":

        mostrar_titulo(
            "info.svg",
            "Acerca de VisionQA",
            "Información general del sistema y las tecnologías utilizadas.",
        )

        st.markdown("""
            **VisionQA** es un sistema inteligente de inspección visual
            desarrollado para apoyar el control de calidad de piezas
            manufacturadas.

            El sistema integra:

            - Visión por computadora.
            - Modelo de detección YOLOv8.
            - Procesamiento de imágenes con OpenCV.
            - Dashboard desarrollado con Streamlit.
            - Análisis de causas mediante Gemini.
            - Principios de Manufactura Esbelta y Six Sigma.
            """)

        st.markdown("### Información del proyecto")

        col1, col2 = st.columns(2)

        with col1:

            st.markdown("""
                **Proyecto:** VisionQA

                **Empresa:** IOT Technologies

                **Área:** Control de Calidad
                """)

        with col2:

            st.markdown("""
                **Desarrolladora:** Dorcas Tabita Perez Martinez

                **Tecnologías:** Python, YOLOv8, OpenCV, Streamlit y Gemini

                **Versión:** 1.0
                """)

        st.divider()

    mostrar_footer()


if __name__ == "__main__":
    main()
